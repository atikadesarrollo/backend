import xmlrpc.client
from flask import Blueprint, jsonify, request, send_file, Response
import base64
from dotenv import load_dotenv
import os
from io import BytesIO, StringIO
from datetime import datetime
import csv
from openpyxl import Workbook
import threading
import tempfile
import uuid

# Crear el Blueprint
odoo_bp = Blueprint('odoo_api', __name__)

# Cargar las variables de entorno desde el archivo .env
load_dotenv()

# Obtener las variables de entorno
url = os.getenv('ODOO_URL')
db = os.getenv('ODOO_DB')
username = os.getenv('ODOO_USERNAME')
password = os.getenv('ODOO_PASSWORD')

EXPORT_JOBS = {}
EXPORT_JOBS_LOCK = threading.Lock()
EXPORT_BASE_DIR = os.path.join(tempfile.gettempdir(), 'atika_odoo_exports')
os.makedirs(EXPORT_BASE_DIR, exist_ok=True)


class OdooServiceClient:
    def __init__(self, models_proxy, uid):
        self.models_proxy = models_proxy
        self.uid = uid

    def execute_kw(self, model, method, args=None, kwargs=None):
        if args is None:
            args = []
        if kwargs is None:
            kwargs = {}

        return self.models_proxy.execute_kw(
            db,
            self.uid,
            password,
            model,
            method,
            args,
            kwargs
        )


def get_odoo_client():
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url), allow_none=True)
    uid = common.authenticate(db, username, password, {})

    if not uid:
        raise Exception("No se pudo autenticar en Odoo")

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url), allow_none=True)
    return OdooServiceClient(models, uid)


def get_sales_analysis_2025_grouped(odoo_client):
    domain = [
        ("np_date_order", ">=", "2025-01-01 00:00:00"),
        ("np_date_order", "<",  "2026-01-01 00:00:00"),
        ("state", "=", "sale"),
    ]

    fields = [
        "np_ov_docnum_ref",
        "price_subtotal:sum",
        "product_uom_qty:sum",
        "qty_delivered:sum",
        "qty_invoiced:sum",
        "__count",
    ]
    groupby = ["np_ov_docnum_ref"]

    rows = odoo_client.execute_kw(
        "sale.order.line",
        "read_group",
        [domain, fields, groupby],
        {"lazy": False}
    )

    result = []
    for row in rows:
        ref_value = row.get("np_ov_docnum_ref")
        ref_name = ref_value[1] if isinstance(ref_value, (list, tuple)) and len(ref_value) > 1 else "Sin referencia"

        result.append({
            "pedido_referencia": ref_name,
            "lineas": row.get("__count", 0),
            "subtotal_sum": row.get("price_subtotal_sum", 0.0),
            "qty_sum": row.get("product_uom_qty_sum", 0.0),
            "qty_delivered_sum": row.get("qty_delivered_sum", 0.0),
            "qty_invoiced_sum": row.get("qty_invoiced_sum", 0.0),
        })

    return {
        "year": 2025,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "total_groups": len(result),
        "data": result,
    }


def get_sales_order_lines_2025_raw(odoo_client):
    domain = [
        ("np_date_order", ">=", "2025-01-01 00:00:00"),
        ("np_date_order", "<", "2026-01-01 00:00:00"),
        ("state", "=", "sale"),
    ]

    rows = odoo_client.execute_kw(
        "sale.order.line",
        "search_read",
        [domain],
        {"limit": False}
    )

    return {
        "year": 2025,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "total_records": len(rows),
        "data": rows,
    }


def build_sales_analysis_2025_grouped_csv(data_rows):
    output = StringIO()
    writer = csv.writer(output)

    if not data_rows:
        return output.getvalue()

    headers = list(data_rows[0].keys())
    writer.writerow(headers)
    for row in data_rows:
        writer.writerow([str(row.get(header, "")) for header in headers])

    return output.getvalue()


def build_sales_analysis_2025_grouped_xlsx(data_rows):
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title='sales_analysis_2025')

    if data_rows:
        headers = list(data_rows[0].keys())
        worksheet.append(headers)
        for row in data_rows:
            worksheet.append([str(row.get(header, "")) for header in headers])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def get_sales_order_lines_2025_raw_batched(odoo_client, limit=None, batch_size=1000):
    domain = [
        ("np_date_order", ">=", "2025-01-01 00:00:00"),
        ("np_date_order", "<", "2026-01-01 00:00:00"),
        ("state", "=", "sale"),
    ]

    fields = [
        "id",
        "order_id",
        "np_ov_docnum_ref",
        "np_date_order",
        "state",
        "product_id",
        "name",
        "product_uom_qty",
        "qty_delivered",
        "qty_invoiced",
        "price_unit",
        "discount",
        "price_subtotal",
    ]

    record_ids = odoo_client.execute_kw(
        "sale.order.line",
        "search",
        [domain],
        {"order": "id asc"}
    )

    if limit and limit > 0:
        record_ids = record_ids[:limit]

    rows = []
    if record_ids:
        for start in range(0, len(record_ids), batch_size):
            batch_ids = record_ids[start:start + batch_size]
            batch_rows = odoo_client.execute_kw(
                "sale.order.line",
                "read",
                [batch_ids],
                {"fields": fields}
            )
            rows.extend(batch_rows)

    return {
        "year": 2025,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "total_records": len(rows),
        "data": rows,
    }


def get_sales_order_lines_raw_batched_by_range(odoo_client, date_start, date_end, limit=None, batch_size=1000):
    domain = [
        ("np_date_order", ">=", date_start),
        ("np_date_order", "<", date_end),
        ("state", "=", "sale"),
    ]

    fields = [
        "id",
        "order_id",
        "np_ov_docnum_ref",
        "np_date_order",
        "state",
        "product_id",
        "name",
        "product_uom_qty",
        "qty_delivered",
        "qty_invoiced",
        "price_unit",
        "discount",
        "price_subtotal",
    ]

    record_ids = odoo_client.execute_kw(
        "sale.order.line",
        "search",
        [domain],
        {"order": "id asc"}
    )

    if limit and limit > 0:
        record_ids = record_ids[:limit]

    rows = []
    if record_ids:
        for start in range(0, len(record_ids), batch_size):
            batch_ids = record_ids[start:start + batch_size]
            batch_rows = odoo_client.execute_kw(
                "sale.order.line",
                "read",
                [batch_ids],
                {"fields": fields}
            )
            rows.extend(batch_rows)

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "total_records": len(rows),
        "data": rows,
    }


def run_december_2025_export_job(job_id, export_format, limit=None, batch_size=1000):
    try:
        with EXPORT_JOBS_LOCK:
            EXPORT_JOBS[job_id]["status"] = "processing"
            EXPORT_JOBS[job_id]["started_at"] = datetime.utcnow().isoformat() + "Z"

        odoo_client = get_odoo_client()
        payload = get_sales_order_lines_raw_batched_by_range(
            odoo_client,
            "2025-12-01 00:00:00",
            "2026-01-01 00:00:00",
            limit=limit,
            batch_size=batch_size
        )

        rows = payload.get("data", [])
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')

        if export_format == 'csv':
            filename = f"sales_order_line_december_2025_{timestamp}.csv"
            file_path = os.path.join(EXPORT_BASE_DIR, f"{job_id}_{filename}")
            csv_content = build_sales_analysis_2025_grouped_csv(rows)
            with open(file_path, 'w', encoding='utf-8-sig', newline='') as file_obj:
                file_obj.write(csv_content)
            mimetype = 'text/csv; charset=utf-8'
        else:
            filename = f"sales_order_line_december_2025_{timestamp}.xlsx"
            file_path = os.path.join(EXPORT_BASE_DIR, f"{job_id}_{filename}")
            xlsx_stream = build_sales_analysis_2025_grouped_xlsx(rows)
            with open(file_path, 'wb') as file_obj:
                file_obj.write(xlsx_stream.getvalue())
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        with EXPORT_JOBS_LOCK:
            EXPORT_JOBS[job_id].update({
                "status": "completed",
                "completed_at": datetime.utcnow().isoformat() + "Z",
                "total_records": payload.get("total_records", 0),
                "file_path": file_path,
                "filename": filename,
                "mimetype": mimetype
            })
    except Exception as exc:
        with EXPORT_JOBS_LOCK:
            EXPORT_JOBS[job_id].update({
                "status": "failed",
                "completed_at": datetime.utcnow().isoformat() + "Z",
                "error": str(exc)
            })


ANALISIS_VENTA_EXPORT_FIELDS = [
    "np_canal",
    "qty_delivered",
    "qty_invoiced",
    "np_sei_cdp",
    "order_partner_id/name",
    "np_sei_cotfin",
    "product_uom_qty",
    "np_departamento",
    "name",
    "discount",
    "np_original_discount",
    "np_docnum_oferta",
    "np_ov_docnum_ref/name",
    "categ_id/name",
    "np_date_order",
    "source_currency_id/name",
    "project_id/name",
    "order_id/client_order_ref",
    "np_rpt_discount",
    "np_rpt_flete_unitario",
    "np_rpt_flete_unitario_mo",
    "np_rpt_base_price",
    "np_rpt_price_unit",
    "np_rpt_subtotal",
    "np_rpt_total",
    "order_partner_id/vat",
    "np_product_sku",
    "np_rate_currency",
    "np_tipo_despacho",
    "np_total_discount_promotion",
    "salesman_id/name",
    "np_venta_anticipada",
    "analytic_line_ids/account_id",
    "analytic_line_ids/amount",
    "analytic_line_ids/project_id",
]

ANALISIS_VENTA_HEADERS = [
    "Canal",
    "Cantidad de entrega",
    "Cantidad Facturada",
    "CDP",
    "Cliente",
    "Cotización Final",
    "Ctd. de producto",
    "Departamento",
    "Descripción",
    "Descuento (%)",
    "Descuento original",
    "Docnum Oferta",
    "Docnum OV",
    "Familia",
    "Fecha de pedido",
    "Moneda origen",
    "Proyecto",
    "Referencia de pedido",
    "RPT Descuento",
    "RPT Flete Unitario",
    "RPT Flete Unitario MO",
    "RPT Precio Base",
    "RPT Precio Unitario",
    "RPT Subtotal",
    "RPT Total",
    "RUT Cliente",
    "SKU",
    "Tasa de Cambio",
    "Tipo Despacho",
    "Total descuento + promoción",
    "Vendedor",
    "Venta Anticipada",
    "Líneas analíticas/Cuenta analítica",
    "Líneas analíticas/Importe",
    "Líneas analíticas/Project",
]


def build_december_2025_xlsx_using_export(odoo_client, batch_size=1000):
    domain = [
        ("np_date_order", ">=", "2025-12-01 00:00:00"),
        ("np_date_order", "<", "2026-01-01 00:00:00"),
        ("state", "=", "sale"),
    ]

    record_ids = odoo_client.execute_kw(
        "sale.order.line",
        "search",
        [domain],
        {"order": "id asc"}
    )

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title='Analisis_Venta_Dic2025')
    worksheet.append(ANALISIS_VENTA_HEADERS)

    for start in range(0, len(record_ids), batch_size):
        batch_ids = record_ids[start:start + batch_size]

        export_result = odoo_client.execute_kw(
            "sale.order.line",
            "export_data",
            [batch_ids, ANALISIS_VENTA_EXPORT_FIELDS],
            {}
        )

        rows = export_result.get("datas", [])
        for row in rows:
            worksheet.append(row)

    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    filename = f"analisis_venta_diciembre_2025_{timestamp}.xlsx"
    file_path = os.path.join(EXPORT_BASE_DIR, f"{uuid.uuid4()}_{filename}")
    workbook.save(file_path)

    return file_path, filename


@odoo_bp.route('/reports/sales-analysis/2025/grouped', methods=['GET'])
def sales_analysis_2025_grouped_endpoint():
    try:
        odoo_client = get_odoo_client()
        file_path, filename = build_december_2025_xlsx_using_export(odoo_client, batch_size=500)
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@odoo_bp.route('/reports/sales-analysis/2025/december/export-async/start', methods=['POST', 'GET'])
def start_december_2025_async_export():
    try:
        export_format = request.args.get('format', 'xlsx').lower()
        if export_format not in ['xlsx', 'csv']:
            return jsonify({"error": "Formato inválido. Usa format=xlsx o format=csv"}), 400

        limit = request.args.get('limit', type=int)
        batch_size = request.args.get('batch_size', default=1000, type=int)
        if not batch_size or batch_size <= 0:
            batch_size = 1000
        if batch_size > 5000:
            batch_size = 5000

        job_id = str(uuid.uuid4())
        with EXPORT_JOBS_LOCK:
            EXPORT_JOBS[job_id] = {
                "job_id": job_id,
                "status": "queued",
                "created_at": datetime.utcnow().isoformat() + "Z",
                "format": export_format,
                "limit": limit,
                "batch_size": batch_size,
                "period": "2025-12"
            }

        worker = threading.Thread(
            target=run_december_2025_export_job,
            args=(job_id, export_format, limit, batch_size),
            daemon=True
        )
        worker.start()

        base_path = '/odoo/reports/sales-analysis/2025/december/export-async'
        return jsonify({
            "message": "Exportación en proceso",
            "job_id": job_id,
            "status": "queued",
            "period": "2025-12",
            "status_url": f"{base_path}/status/{job_id}",
            "download_url": f"{base_path}/download/{job_id}"
        }), 202
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@odoo_bp.route('/reports/sales-analysis/2025/december/export-async/status/<string:job_id>', methods=['GET'])
def get_december_2025_async_export_status(job_id):
    with EXPORT_JOBS_LOCK:
        job = EXPORT_JOBS.get(job_id)

    if not job:
        return jsonify({"error": "Job no encontrado"}), 404

    response = {
        "job_id": job.get("job_id"),
        "status": job.get("status"),
        "period": job.get("period"),
        "format": job.get("format"),
        "created_at": job.get("created_at"),
        "started_at": job.get("started_at"),
        "completed_at": job.get("completed_at"),
        "total_records": job.get("total_records")
    }

    if job.get("status") == "failed":
        response["error"] = job.get("error")

    if job.get("status") == "completed":
        response["download_url"] = f"/odoo/reports/sales-analysis/2025/december/export-async/download/{job_id}"

    return jsonify(response), 200


@odoo_bp.route('/reports/sales-analysis/2025/december/export-async/download/<string:job_id>', methods=['GET'])
def download_december_2025_async_export(job_id):
    with EXPORT_JOBS_LOCK:
        job = EXPORT_JOBS.get(job_id)

    if not job:
        return jsonify({"error": "Job no encontrado"}), 404

    if job.get("status") != "completed":
        return jsonify({
            "error": "El archivo aún no está disponible",
            "status": job.get("status")
        }), 409

    file_path = job.get("file_path")
    if not file_path or not os.path.exists(file_path):
        return jsonify({"error": "Archivo no encontrado en servidor"}), 404

    return send_file(
        file_path,
        mimetype=job.get("mimetype"),
        as_attachment=True,
        download_name=job.get("filename")
    )

@odoo_bp.route('/sale_orders/date/<string:date_range>', methods=['GET'])
@odoo_bp.route('/sale_orders/<string:order_names>', methods=['GET'])
def get_sale_orders(date_range=None, order_names=None):
    try:
        domain = []
        if date_range:
            params = date_range.split('&')
            if len(params) != 2 or not all(len(p) == 10 and p.count('-') == 2 for p in params):
                return jsonify({"error": "Para búsqueda por fechas use: YYYY-MM-DD&YYYY-MM-DD"}), 400
            domain = [
                ['date_order', '>=', params[0]],
                ['date_order', '<=', params[1]]
            ]
        elif order_names:
            params = order_names.split('&')
            domain = [['name', 'in', params]]
        else:
            return jsonify({"error": "Parámetros de búsqueda no proporcionados"}), 400

        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

        sale_order_ids = models.execute_kw(db, uid, password,
            'sale.order', 'search',
            [domain])

        if sale_order_ids:
            header_fields = [
                'name', 'state', 'project_id', 'construction_company', 'decorator', 
                'architect', 'partner_id', 'partner_invoice_id', 
                'partner_shipping_id', 'client_order_ref', 'payment_term_id','date_order', 
                'validity_date', 'note','np_sei_cotfin', 'np_sei_vreserva', 'np_sei_vcalzada', 
                'np_sei_factanticpag', 'np_sei_esp', 'np_sei_arq2', 'np_sei_tdes',
                'np_sei_einm', 'np_sei_obc', 'np_sei_pesodoc', 'np_sei_nob', 'np_oport', 
                'np_sei_autoriztc', 'np_origen_documento', 'order_line', 'intermediary', 'nxt_id_erp',
                'user_id','ov_nxt_sync', 'ov_nxt_id_erp', 'ov_docnum_ref'
            ]

            sale_orders = models.execute_kw(db, uid, password,
                'sale.order', 'read',
                [sale_order_ids],
                {'fields': header_fields})

            # Obtener información de empleados
            for order in sale_orders:
                if order.get('user_id'):
                    employee_ids = models.execute_kw(db, uid, password,
                        'hr.employee', 'search',
                        [[('user_id', '=', order['user_id'][0])]])
                    
                    if employee_ids:
                        employee_data = models.execute_kw(db, uid, password,
                            'hr.employee', 'read',
                            [employee_ids[0]],
                            {'fields': ['np_sei_geo', 'np_sei_une', 'np_sei_area', 'code']})
                        
                        if employee_data:
                            order['np_sei_geo'] = employee_data[0].get('np_sei_geo', False)
                            order['np_sei_une'] = employee_data[0].get('np_sei_une', False)
                            order['np_sei_area'] = employee_data[0].get('np_sei_area', False)
                            order['employee_code'] = employee_data[0].get('code', False)
            # Obtener los códigos de nxt_id_erp
            partner_ids = [order['partner_id'][0] for order in sale_orders if order['partner_id']]
            partners = models.execute_kw(db, uid, password,
                'res.partner', 'read',
                [partner_ids],
                {'fields': ['nxt_id_erp']})

            partner_dict = {partner['id']: partner['nxt_id_erp'] for partner in partners}

            for order in sale_orders:
                partner_id = order.get('partner_id')
                if partner_id:
                    order['nxt_id_erp'] = partner_dict.get(partner_id[0], '')

            line_fields = [
                'product_id', 'product_uom_qty', 'np_rpt_base_price', 'order_id',
                'np_fecha_de_entrega', 'np_rpt_base_price', 'np_mo_price_unit', 'source_currency_id',
                'np_rate_currency', 'discount', 'tax_id', 'price_subtotal', 
                'discount', 'np_original_discount',
                'np_rpt_flete_unitario', 'np_product_sku', 'source_currency_id', 'np_sei_cdp', 'np_has_discount', 'np_rpt_price_unit', 'price_unit', 'freight_cost_origin_currency'

            ]

            order_line_ids = [line_id for order in sale_orders for line_id in order['order_line']]
            sale_order_lines = models.execute_kw(db, uid, password,
                'sale.order.line', 'read',
                [order_line_ids],
                {'fields': line_fields})

            sale_order_lines_grouped = {}
            for line in sale_order_lines:
                order_id = line['order_id'][0]
                if order_id not in sale_order_lines_grouped:
                    sale_order_lines_grouped[order_id] = []
                sale_order_lines_grouped[order_id].append(line)

            for order in sale_orders:
                order['sale_order_line_data'] = sale_order_lines_grouped.get(order['id'], [])

            result = [{"sale_order_data": order} for order in sale_orders]

            return jsonify(result)
        else:
            return jsonify({"error": "No se encontraron órdenes de venta"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@odoo_bp.route('/update_fields/<order_names>', methods=['GET'])
@odoo_bp.route('/update_fields/date/<date_range>', methods=['GET'])
def update_fields(order_names=None, date_range=None):
    try:
        domain = []

        if date_range:
            date_start, date_end = date_range.split('&')
            domain.append(['date_order', '>=', date_start])
            domain.append(['date_order', '<=', date_end])
        elif order_names:
            order_names_list = order_names.split('&')
            domain.append(['name', 'in', order_names_list])
        else:
            return jsonify({"error": "No se proporcionaron parámetros de búsqueda"}), 400

        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
        uid = common.authenticate(db, username, password, {})

        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

        # Buscar las Órdenes de Venta
        sale_order_ids = models.execute_kw(db, uid, password,
            'sale.order', 'search',
            [domain])

        if sale_order_ids:
            # Campos específicos para la cabecera y del modelo sale.order y sale.order.line
            header_fields = [
                'name', 'project_id', 'state', 'construction_company', 'decorator', 
                'architect', 'construction_company', 'order_line', 'intermediary'
            ]

            line_fields = [
                'order_id', 'discount', 'np_original_discount', 'np_rpt_flete_unitario'
            ]

            # Leer los campos de las Órdenes de Venta
            sale_orders = models.execute_kw(db, uid, password,
                'sale.order', 'read',
                [sale_order_ids],
                {'fields': header_fields})

            # Obtener las líneas de las Órdenes de Venta
            order_line_ids = [line_id for order in sale_orders for line_id in order['order_line']]
            sale_order_lines = models.execute_kw(db, uid, password,
                'sale.order.line', 'read',
                [order_line_ids],
                {'fields': line_fields})

            # Ordenar y agrupar las órdenes de venta por el código de la orden
            sale_orders_sorted = sorted(sale_orders, key=lambda x: x['name'])
            sale_order_lines_grouped = {}
            for line in sale_order_lines:
                order_id = line['order_id'][0]
                if order_id not in sale_order_lines_grouped:
                    sale_order_lines_grouped[order_id] = []
                sale_order_lines_grouped[order_id].append(line)

            response = {
                "sale_order_data": sale_orders_sorted,
                "sale_order_line_data": sale_order_lines
            }

            return jsonify(response)
        else:
            return jsonify({"error": "No se encontraron órdenes de venta"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@odoo_bp.route('/sale_orders_status/<string:search_params>', methods=['GET'])
def get_sales_order_status(search_params):
    try:
        domain = []
        params = search_params.split('&')
        
        # Detectar si son fechas (si el primer parámetro tiene formato YYYY-MM-DD)
        if len(params[0]) == 10 and params[0].count('-') == 2:
            if len(params) != 2:
                return jsonify({"error": "Para búsqueda por fechas use: YYYY-MM-DD&YYYY-MM-DD"}), 400
            domain = [
                ('date_order', '>=', params[0]),
                ('date_order', '<=', params[1])
            ]
        # Si no son fechas, buscar por códigos
        else:
            domain = [('name', 'in', params)]

        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

        sale_order_ids = models.execute_kw(db, uid, password,
            'sale.order', 'search',
            [domain])

        if sale_order_ids:
            sale_orders = models.execute_kw(db, uid, password,
                'sale.order', 'read',
                [sale_order_ids],
                {'fields': ['name', 'state']})

            return jsonify({"orders_status": sale_orders})
        else:
            return jsonify({"error": "No sale orders found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500



#endpoint para actualización de campos
@odoo_bp.route('/sale_orders_status2/<string:search_params>', methods=['GET'])
def get_sales_order_status2(search_params):
    try:
        domain = []
        params = search_params.split('&')
        
        # Detectar si son fechas (si el primer parámetro tiene formato YYYY-MM-DD)
        if len(params[0]) == 10 and params[0].count('-') == 2:
            if len(params) != 2:
                return jsonify({"error": "Para búsqueda por fechas use: YYYY-MM-DD&YYYY-MM-DD"}), 400
            # Agrega horas para cubrir todo el rango del día al que corresponda. 
            domain = [
                ('write_date', '>=', f"{params[0]} 00:00:00"),
                ('write_date', '<=', f"{params[1]} 23:59:59")
            ]
        # Si no son fechas, buscar por códigos
        else:
            domain = [('name', 'in', params)]


        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

        sale_order_ids = models.execute_kw(db, uid, password,
            'sale.order', 'search',
            [domain])

    

        if sale_order_ids:
            sale_orders = models.execute_kw(db, uid, password,
                'sale.order', 'read',
                [sale_order_ids],
                {'fields': ['name', 'state', 'write_date', 'date_order', 'create_date', 'validity_date']})


            return jsonify({"orders_status": sale_orders})
        else:
            
            return jsonify({"error": "No se encontraron órdenes de venta"}), 404
    except Exception as e:
       
        return jsonify({"error": str(e)}), 500


@odoo_bp.route('/analisis_proyectos/<string:date_range>', methods=['GET'])
def get_projects_info(date_range):
    try:
        # Validar y dividir el rango de fechas
        dates = date_range.split('&')
        if len(dates) != 2 or not all(len(date) == 10 and date.count('-') == 2 for date in dates):
            return jsonify({"error": "Para búsqueda por fechas usa: YYYY-MM-DD&YYYY-MM-DD"}), 400

        date_start, date_end = dates

        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

        # Busca las órdenes de venta que tienen un project_id asociado y fecha dentro del rango
        sale_order_ids = models.execute_kw(db, uid, password,
            'sale.order', 'search',
            [[('project_id', '!=', False), ('date_order', '>=', date_start), ('date_order', '<=', date_end)]])

        if sale_order_ids:
            sale_orders = models.execute_kw(db, uid, password,
                'sale.order', 'read',
                [sale_order_ids],
                {'fields': ['name', 'state', 'project_id', 'amount_total', 'user_id']})

            # Se obtienen los campos adicionales del modelo custom.projects para cada project_id
            project_ids = [order['project_id'][0] for order in sale_orders if order['project_id']]
            projects = models.execute_kw(db, uid, password,
                'custom.projects', 'read',
                [project_ids],
                {'fields': [
                    'principal', 'project_name', 'project_creator', 'real_estate_specifier', 
                    'arch_specifier', 'start_date', 'closing_date', 'np_canal', 
                    'custom_project_type', 'contract_value', 'project_type', 
                    'business_value', 'sale_status', 'foreclosed_seller', 
                    'decorator', 'architect', 'construction_company','np_canal','name', 'create_date'
                ]})

            project_dict = {project['id']: project for project in projects}

            for order in sale_orders:
                project_id = order.get('project_id')
                if project_id:
                    project_data = project_dict.get(project_id[0], {})
                    order.update({
                        'Mandante': project_data.get('principal',[])[1] if project_data.get('principal') else '',
                        'Nombre_Proyecto': project_data.get('project_name', ''),
                        'Creador_Proyecto': project_data.get('project_creator', [])[1] if project_data.get('project_creator') else '',
                        'Especificador_Inmobiliaria': project_data.get('real_estate_specifier', [])[1] if project_data.get('real_estate_specifier') else '',
                        'Especificador_Arquitectura': project_data.get('arch_specifier', [])[1] if project_data.get('arch_specifier') else '',                       
                        'Fecha_Inicio': project_data.get('start_date', ''),
                        'Fecha_Termino': project_data.get('closing_date', ''),
                        'Canal_Venta': project_data.get('np_canal', ''),
                        'Tipo_Proyecto': project_data.get('custom_project_type', ''),
                        # los campos Tipo_Obra (project_type) son:
                        #('1','Habitacional - Condominio'),
                        #('2','Habitacional - Edificio'),
                        #('3','Locales individuales'),
                        #('4','Oficinas'),
                        #('5','Institucional'),
                        #('6','Hoteles'),
                        #('7','Industria'),
                        #('8','Salud'),
                        #('9','Educacional'),
                        #('10','Casa Particular'),
                        #('11','Cadena de locales'),
                        #('12','Mall'),
                        #('13','Museo-salasexhibición-galería'),
                        #('14','Casa Particular - Obra Nueva'),
                        #('15','Casa Particular - Remodelación'),
                        #('16','SPA'),
                        #('17','Restaurant')
                        'Valor_Contrato': project_data.get('contract_value', ''),
                        'Tipo_Obra': project_data.get('project_type', ''),
                        'Valor_Negocio': project_data.get('business_value', ''),
                        # Los campos Valor_Negocio (business_value) son:
                        #('1','Mayor o igual 200MM$'),
                        #('2','[150;200] MM$'),
                        #('3','[100;150] MM$'),
                        #('4','[60;100] MM$'),
                        #('5','[20;60] MM$'),
                        #('6','[10;20] MM$'),
                        #('7','[5;10] MM$'),
                        #('8','[1;5] MM$'),
                        #('9','[0;1] MM$'),
                        #('10','[1,1;2] MM$'),
                        #('11','[3,1;4] MM$'),
                        #('12','[4,1;5] MM$'),
                        
                        'Estatus_Venta': project_data.get('sale_status', ''),
                       
                        # los campos Estatus_Venta (sale_status) son:
                        #('1','En seguimiento'),
                        #('2','O/C recibida'),
                        #('3','O/C recibida- proceso despacho'),
                        #('4','Despachos finalizados'),
                        #('5','Perdido'),
                        #('6','Traspasado'),
                        #('7','Detectado'),
                        
                        'Vendedor_Adjudicado': project_data.get('foreclosed_seller', [])[1] if project_data.get('foreclosed_seller') else '',
                        'Interiorista': project_data.get('decorator', [])[1] if project_data.get('decorator') else '',
                        'Arquitecto': project_data.get('architect', [])[1] if project_data.get('architect') else '',
                        'Constructora': project_data.get('construction_company', [])[1] if project_data.get('construction_company') else '',
                        'Canal_Venta': project_data.get('np_canal', ''),
                        'Estado_Orden': order.get('state', ''),
                        'Orden_Odoo': order.get('name', ''),
                        'Monto_Total': order.get('amount_total', ''),
                        'Nombre_Proyecto': project_data.get('project_name', ''),
                        'Codigo_Proyecto': project_data.get('name', ''),
                        'Fecha_Creación_Proyecto': project_data.get('create_date', ''),
                        'Vendedor_Pedido_Odoo': order.get('user_id', [])[1] if order.get('user_id') else ''
                    })

            return jsonify({"Pedidos Odoo con Proyecto asociado": sale_orders})
        else:
            return jsonify({"error": "No se encontraron órdenes de venta con project_id"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@odoo_bp.route('/analisis_proyectos2', methods=['GET'])
@odoo_bp.route('/analisis_proyectos2/<string:date_range>', methods=['GET'])
def get_latest_projects_info(date_range=None):
    try:
        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url), allow_none=True)
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url), allow_none=True)

        domain = []
        limit = None
        if date_range:
            dates = date_range.split('&')
            if len(dates) != 2 or not all(len(date) == 10 and date.count('-') == 2 for date in dates):
                return jsonify({"error": "Para búsqueda por fechas usa: YYYY-MM-DD&YYYY-MM-DD"}), 400
            date_start, date_end = dates
            domain = [('create_date', '>=', date_start), ('create_date', '<=', date_end)]
        else:
            # Si no se especifican fechas, busca los últimos 50 proyectos creados 
            limit = 50

        # Busca todos los proyectos existentes y ordena por fecha de creación descendente
        projects = models.execute_kw(db, uid, password,
            'custom.projects', 'search_read',
            [domain],
            {'fields': [
                'principal', 'project_name', 'project_creator', 'real_estate_specifier', 
                'arch_specifier', 'start_date', 'closing_date', 'np_canal',
                'custom_project_type', 'contract_value', 'project_type', 
                'business_value', 'sale_status', 'foreclosed_seller', 
                'decorator', 'architect', 'construction_company','np_canal','name', 'state_id', 'contract_balance', 'create_date' 
            ], 'order': 'create_date desc', 'limit': limit})

        if projects:
            # Obtener todos los user_ids de los project_creators
            user_ids = [project['project_creator'][0] for project in projects if project.get('project_creator')]

            if user_ids:
                # Obtener los department_id de todos los project_creators en una sola llamada
                employees = models.execute_kw(db, uid, password,
                    'hr.employee', 'search_read',
                    [[('user_id', 'in', user_ids)]],
                    {'fields': ['user_id', 'department_id', 'np_sei_geo']})
                # obtiuene los emails (login) de los project_creators
                users = models.execute_kw(db, uid, password,
                    'res.users', 'read',
                    [user_ids],
                    {'fields': ['login']})  
                
                # Crear un diccionario para mapear user_id a department_id y los nuevos campos del modelo hr.employee
                user_to_department = {employee['user_id'][0]: employee['department_id'][1] if employee['department_id'] else '' for employee in employees}
                user_to_email = {user['id']: user['login'] for user in users}
                user_to_geo = {employee['user_id'][0]: employee.get('np_sei_geo', '') for employee in employees}
                
                # Asignar el department_id a cada proyecto
                for project in projects:
                    project_creator = project.get('project_creator')
                    if project_creator:
                        project['department_id'] = user_to_department.get(project_creator[0], '')
                        project['creator_email'] = user_to_email.get(project_creator[0], '')
                        project['np_sei_geo'] = user_to_geo.get(project_creator[0], '')
            
            # Obtener los work_emails de los arch_specifier
            arch_specifier_ids = []
            for project in projects:
                if project.get('arch_specifier'):
                    arch_specifier_ids.append(project['arch_specifier'][0])
            
            employee_to_work_email = {}
            if arch_specifier_ids:
                arch_employees = models.execute_kw(db, uid, password,
                    'hr.employee', 'read',
                    [arch_specifier_ids],
                    {'fields': ['work_email']})
                
                employee_to_work_email = {employee['id']: employee.get('work_email', '') for employee in arch_employees}
            
            # Asignar el work_email a cada proyecto
            for project in projects:
                if project.get('arch_specifier'):
                    project['arch_specifier_email'] = employee_to_work_email.get(project['arch_specifier'][0], '')
                else:
                    project['arch_specifier_email'] = ''
            
            # Obtener los work_emails de los real_estate_specifier
            real_estate_specifier_ids = []
            for project in projects:
                if project.get('real_estate_specifier'):
                    real_estate_specifier_ids.append(project['real_estate_specifier'][0])
            
            real_estate_employee_to_work_email = {}
            if real_estate_specifier_ids:
                real_estate_employees = models.execute_kw(db, uid, password,
                    'hr.employee', 'read',
                    [real_estate_specifier_ids],
                    {'fields': ['work_email']})
                
                real_estate_employee_to_work_email = {employee['id']: employee.get('work_email', '') for employee in real_estate_employees}
            
            # Asignar el work_email del real_estate_specifier a cada proyecto
            for project in projects:
                if project.get('real_estate_specifier'):
                    project['real_estate_specifier_email'] = real_estate_employee_to_work_email.get(project['real_estate_specifier'][0], '')
                else:
                    project['real_estate_specifier_email'] = ''
            
            # Obtener las órdenes de venta asociadas a los proyectos
            project_ids = [project['id'] for project in projects]
            sale_orders = models.execute_kw(db, uid, password,
                'sale.order', 'search_read',
                [[('project_id', 'in', project_ids), ('state', '!=', 'cancel')]],
                {'fields': ['name', 'state', 'project_id', 'amount_total', 'user_id']})
            
            # Obtener los department_id de los user_id asociados a las sale_orders
            sale_order_user_ids = [order['user_id'][0] for order in sale_orders if order.get('user_id')]
            if sale_order_user_ids:
                sale_order_employees = models.execute_kw(db, uid, password,
                    'hr.employee', 'search_read',
                    [[('user_id', 'in', sale_order_user_ids)]],
                    {'fields': ['user_id', 'department_id']})

                # Crear un diccionario para mapear user_id a department_id
                sale_order_user_to_department = {employee['user_id'][0]: employee['department_id'][1] if employee['department_id'] else '' for employee in sale_order_employees}

                # Asignar el department_id a cada orden de venta
                for order in sale_orders:
                    user_id = order.get('user_id')
                    if user_id:
                        order['department_id'] = sale_order_user_to_department.get(user_id[0], '')

            # Crear un diccionario para mapear project_id a sus órdenes de venta
            project_to_sale_orders = {}
            for order in sale_orders:
                project_id = order['project_id'][0]
                if project_id not in project_to_sale_orders:
                    project_to_sale_orders[project_id] = []
                # Decorar los datos internos de las órdenes de venta
                formatted_order = {
                    'Orden': order.get('name', ''),
                    'Estado': order.get('state', ''),
                    'Codigo_Proyecto': order.get('project_id', [])[1] if order.get('project_id') else '',
                    'Monto_Total': order.get('amount_total', ''),
                    'Usuario': order.get('user_id', [])[1] if order.get('user_id') else '',
                    'Departamento_Vendedor': order.get('department_id', '')

                }
                project_to_sale_orders[project_id].append(formatted_order)

            # Decorar el resultado con los campos formateados
            formatted_projects = []
            for project in projects:
                formatted_project = {
                    'Mandante': project.get('principal', [])[1] if project.get('principal') else '',
                    'Nombre_Proyecto': project.get('project_name', ''),
                    'Creador_Proyecto': project.get('project_creator', [])[1] if project.get('project_creator') else '',
                    'Especificador_Inmobiliaria': project.get('real_estate_specifier', [])[1] if project.get('real_estate_specifier') else '',
                    'Email_Especificador_Inmobiliaria': project.get('real_estate_specifier_email', ''),
                    'Especificador_Arquitectura': project.get('arch_specifier', [])[1] if project.get('arch_specifier') else '',
                    'Email_Especificador_Arquitectura': project.get('arch_specifier_email', ''),
                    'Fecha_Inicio': project.get('start_date', ''),
                    'Fecha_Termino': project.get('closing_date', ''),
                    'Canal_Venta': project.pop('np_canal', ''),
                    'Tipo_Proyecto': project.get('custom_project_type', ''),
                    'Valor_Contrato': project.get('contract_value', ''),
                    'Tipo_Obra': project.get('project_type', ''),
                    'Valor_Negocio': project.get('business_value', ''),
                    'Estatus_Venta': project.get('sale_status', ''),
                    'Vendedor_Adjudicado': project.get('foreclosed_seller', [])[1] if project.get('foreclosed_seller') else '',
                    'Interiorista': project.get('decorator', [])[1] if project.get('decorator') else '',
                    'Arquitecto': project.get('architect', [])[1] if project.get('architect') else '',
                    'Constructora': project.get('construction_company', [])[1] if project.get('construction_company') else '',
                    'Codigo_Proyecto': project.get('name', ''),
                    'Departamento': project.get('department_id', ''),
                    'Ordenes_de_Venta': project_to_sale_orders.get(project['id'], []),
                    'Region': project.get('state_id', [])[1] if project.get('state_id') else '',
                    'Saldo_Contrato': project.get('contract_balance', ''),
                    'Fecha_Creación': project.get('create_date', ''),
                    'Email_Creador': project.get('creator_email', ''),
                    'Geo_Creador': project.get('np_sei_geo', '')
                }
                formatted_projects.append(formatted_project)

            return jsonify({"projects": formatted_projects})
        else:
            return jsonify({"error": "No se encontraron proyectos"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@odoo_bp.route('/ultimo_proyecto', methods=['GET'])
def obtener_ultimo_proyecto ():
    try:
        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url), allow_none=True)
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url), allow_none=True)

        # Busca el último proyecto creado
        project_ids = models.execute_kw(db, uid, password,
            'custom.projects', 'search',
            [[]],  
            {'limit': 1, 'order': 'create_date desc'})

        if project_ids:
            projects = models.execute_kw(db, uid, password,
                'custom.projects', 'read',
                [project_ids],
                {'fields': ['name', 'create_date']})

            return jsonify({"ultimo_proyecto": projects[0]})
        else:
            return jsonify({"error": "No se encontraron proyectos"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500
@odoo_bp.route('/DescargaOC/<string:order_name>', methods=['GET'])
def descarga_oc(order_name):
    try:
        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url), allow_none=True)
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url), allow_none=True)

        # Buscar la orden de venta por nombre
        sale_order_ids = models.execute_kw(db, uid, password,
            'sale.order', 'search',
            [[('name', '=', order_name)]],
            {'limit': 1})

        if not sale_order_ids:
            return jsonify({"error": "Orden de venta no encontrada"}), 404

        # Leer el campo np_oc_file
        sale_orders = models.execute_kw(db, uid, password,
            'sale.order', 'read',
            [sale_order_ids],
            {'fields': ['np_oc_file']})

        np_oc_file = sale_orders[0].get('np_oc_file')
        if not np_oc_file:
            return jsonify({"error": "No existe archivo OC para esta orden"}), 404

        # Decodificar el archivo base64
        pdf_data = base64.b64decode(np_oc_file)
        pdf_stream = BytesIO(pdf_data)

        return send_file(
            pdf_stream,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'OC_{order_name}.pdf'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@odoo_bp.route('/factura/<string:factura_name>', methods=['GET'])
def obtener_factura(factura_name):
    try:
        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url), allow_none=True)
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url), allow_none=True)

        # Buscar la factura por su nombre (número)
        factura_ids = models.execute_kw(db, uid, password,
            'account.move', 'search',
            [[('name', '=', factura_name), ('move_type', 'in', ['out_invoice', 'out_refund'])]], 
            {'limit': 1})

        if not factura_ids:
            return jsonify({"error": "Factura no encontrada"}), 404

        # Obtener todos los campos disponibles del modelo account.move
        fields_data = models.execute_kw(db, uid, password, 
            'account.move', 'fields_get', 
            [], {'attributes': ['string', 'type', 'required']})
        
        # Extraer solo los nombres de los campos
        all_fields = list(fields_data.keys())

        # Obtener datos completos de la factura con todos los campos
        factura = models.execute_kw(db, uid, password,
            'account.move', 'read',
            [factura_ids[0]],
            {'fields': all_fields})

        if not factura:
            return jsonify({"error": "Error al leer la factura"}), 500

        # Obtener todos los campos disponibles de las líneas de factura
        line_fields_data = models.execute_kw(db, uid, password, 
            'account.move.line', 'fields_get', 
            [], {'attributes': ['string', 'type', 'required']})
        
        line_all_fields = list(line_fields_data.keys())

        # Obtener datos de las líneas de la factura
        line_ids = factura[0].get('invoice_line_ids', [])
        lines = []
        
        if line_ids:
            lines = models.execute_kw(db, uid, password,
                'account.move.line', 'read',
                [line_ids],
                {'fields': line_all_fields})

        # Obtener información del cliente (partner)
        partner_id = factura[0].get('partner_id', [0])[0]
        partner_info = {}
        
        if partner_id:
            partner_fields_data = models.execute_kw(db, uid, password, 
                'res.partner', 'fields_get', 
                [], {'attributes': ['string', 'type', 'required']})
            
            partner_all_fields = list(partner_fields_data.keys())
            
            partner_data = models.execute_kw(db, uid, password,
                'res.partner', 'read',
                [partner_id],
                {'fields': partner_all_fields})
                
            if partner_data:
                partner_info = partner_data[0]

        # Preparar la respuesta
        response = {
            "factura": factura[0],
            "lineas": lines,
            "cliente": partner_info
        }

        return jsonify(response)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@odoo_bp.route('/factura_refoliar/<string:factura_name>', methods=['GET'])
def get_factura_refoliar(factura_name):
    try:
        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url), allow_none=True)
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url), allow_none=True)


        # Buscar la factura por su nombre (número)
        factura_ids = models.execute_kw(db, uid, password,
            'account.move', 'search',
            [[('name', '=', factura_name), ('move_type', 'in', ['out_invoice', 'out_refund'])]], 
            {'limit': 1})

        if not factura_ids:
            return jsonify({"error": "Factura no encontrada"}), 404

        # Obtener los campos específicos solicitados
        factura = models.execute_kw(db, uid, password,
            'account.move', 'read',
            [factura_ids[0]],
            {'fields': ['name', 'folio_dte', 'febos_id', 'mensaje_febos', 'seguimiento_febos', 'codigo_febos', 'errores_febos', 'internal_id_febos', 'display_name']})

        if not factura:
            return jsonify({"error": "Error al leer la factura"}), 500

        # Preparar la respuesta con los campos específicos
        response = {
            "Codigo Odoo Factura": factura[0].get('name', ''),
            "Folio DTE": factura[0].get('folio_dte', ''),
            "ID Febos": factura[0].get('febos_id', ''),
            "Mensaje Febos": factura[0].get('mensaje_febos', ''),
            "Seguimiento Febos": factura[0].get('seguimiento_febos', ''),
            "Codigo Febos": factura[0].get('codigo_febos', ''),
            "Errores Febos": factura[0].get('errores_febos', ''),
            "Internal ID Febos": factura[0].get('internal_id_febos', ''),
            "Display Name": factura[0].get('display_name', '')
        }

        return jsonify(response)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@odoo_bp.route('/factura_refoliar/<string:factura_name>', methods=['PUT'])
def update_factura_refoliar(factura_name):
    try:
        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url), allow_none=True)
        uid = common.authenticate(db, username, password, {})
        models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url), allow_none=True)

        # Buscar la factura por su nombre (número)
        factura_ids = models.execute_kw(db, uid, password,
            'account.move', 'search',
            [[('name', '=', factura_name), ('move_type', 'in', ['out_invoice', 'out_refund'])]], 
            {'limit': 1})

        if not factura_ids:
            return jsonify({"error": "Factura no encontrada"}), 404

        # Obtener los datos del request JSON
        data = request.get_json()
        if not data:
            return jsonify({"error": "No se proporcionaron datos para actualizar"}), 400

        # Mapear los campos del JSON a los campos de Odoo
        update_values = {}
        field_mapping = {
            'folio_dte': 'folio_dte',
            'febos_id': 'febos_id', 
            'mensaje_febos': 'mensaje_febos',
            'seguimiento_febos': 'seguimiento_febos',
            'codigo_febos': 'codigo_febos',
            'errores_febos': 'errores_febos',
            'internal_id_febos': 'internal_id_febos',
            'display_name': 'display_name'
        }

        # Construir el diccionario de valores a actualizar
        for json_field, odoo_field in field_mapping.items():
            if json_field in data:
                update_values[odoo_field] = data[json_field]

        if not update_values:
            return jsonify({"error": "No se encontraron campos válidos para actualizar"}), 400

        # Actualizar la factura
        result = models.execute_kw(db, uid, password,
            'account.move', 'write',
            [factura_ids, update_values])

        if result:
            # Obtener los datos actualizados para confirmar
            updated_factura = models.execute_kw(db, uid, password,
                'account.move', 'read',
                [factura_ids[0]],
                {'fields': ['name', 'folio_dte', 'febos_id', 'mensaje_febos', 'seguimiento_febos', 'codigo_febos', 'errores_febos', 'internal_id_febos', 'display_name']})

            response = {
                "message": "Factura actualizada exitosamente",
                "factura_actualizada": {
                    "Codigo Odoo Factura": updated_factura[0].get('name', ''),
                    "Folio DTE": updated_factura[0].get('folio_dte', ''),
                    "ID Febos": updated_factura[0].get('febos_id', ''),
                    "Mensaje Febos": updated_factura[0].get('mensaje_febos', ''),
                    "Seguimiento Febos": updated_factura[0].get('seguimiento_febos', ''),
                    "Codigo Febos": updated_factura[0].get('codigo_febos', ''),
                    "Errores Febos": updated_factura[0].get('errores_febos', ''),
                    "Internal ID Febos": updated_factura[0].get('internal_id_febos', ''),
                    "Display Name": updated_factura[0].get('display_name', '') 
                }
            }
            return jsonify(response)
        else:
            return jsonify({"error": "Error al actualizar la factura"}), 500

    except Exception as e:
        return jsonify({"error": str(e)}), 500

